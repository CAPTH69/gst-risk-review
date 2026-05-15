"""
report_writer.py - Generate a CA-readable Excel exception report.

Takes the enriched reconciliation DataFrame (after risk scoring) and:
  1. Filters to rows that need CA attention (exceptions only)
  2. Calculates ITC at risk per row
  3. Adds a suggested CA action
  4. Writes a formatted Excel file to the reports/ folder

The output is designed to be opened directly by the CA firm. Each row
is colour-coded by risk level (red = High, amber = Medium).
"""

import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill

# Colours for row highlighting (Excel hex, no '#')
HIGH_FILL = PatternFill(fgColor="FFCCCC", fill_type="solid")
MEDIUM_FILL = PatternFill(fgColor="FFE699", fill_type="solid")

# Mapping from internal column names → CA-readable header names
COLUMN_RENAME = {
    "invoice_no": "Invoice No",
    "supplier_gstin": "Supplier GSTIN",
    "supplier_name": "Supplier Name",
    "status": "Reconciliation Status",
    "mismatch_reason": "Mismatch Reason",
    "supplier_risk_level": "Supplier Risk Level",
    "supplier_risk_score": "Supplier Risk Score",
    "supplier_risk_reasons": "Supplier Risk Reasons",
    "purchase_total_itc": "Purchase ITC",
    "gstr2b_total_itc": "GSTR-2B ITC",
    "itc_at_risk": "ITC At Risk",
    "suggested_ca_action": "Suggested CA Action",
}

# Approximate column widths (characters) for each CA-readable column
COLUMN_WIDTHS = {
    "Invoice No": 18,
    "Supplier GSTIN": 20,
    "Supplier Name": 25,
    "Reconciliation Status": 22,
    "Mismatch Reason": 35,
    "Supplier Risk Level": 20,
    "Supplier Risk Score": 20,
    "Supplier Risk Reasons": 40,
    "Purchase ITC": 15,
    "GSTR-2B ITC": 15,
    "ITC At Risk": 15,
    "Suggested CA Action": 35,
    "CA Review Status": 20,
    "CA Remarks": 35,
}


def calculate_itc_at_risk(row):
    """
    Return the rupee amount of ITC at risk for this invoice row.

    The amount depends on what kind of reconciliation issue was found:
    - missing_in_2b: the entire purchase ITC is at risk (supplier never reported it)
    - amount_mismatch: only the difference is disputed
    - duplicate_in_purchase: the duplicate ITC claim is at risk
    - extra_in_2b: no ITC risk (this is in GSTR-2B, not claimed in PR)
    - matched but High/Medium risk supplier: ITC may be denied due to supplier status
    - everything else: 0
    """
    status = row.get("status", "")
    risk_level = row.get("supplier_risk_level", "Low")

    pr_itc = row.get("purchase_total_itc", 0.0)
    g2b_itc = row.get("gstr2b_total_itc", 0.0)

    # Safely convert to float (handle NaN)
    try:
        pr_itc = float(pr_itc) if pd.notna(pr_itc) else 0.0
    except (TypeError, ValueError):
        pr_itc = 0.0

    try:
        g2b_itc = float(g2b_itc) if pd.notna(g2b_itc) else 0.0
    except (TypeError, ValueError):
        g2b_itc = 0.0

    if status == "missing_in_2b":
        return pr_itc

    if status == "amount_mismatch":
        return abs(pr_itc - g2b_itc)

    if status == "duplicate_in_purchase":
        return pr_itc

    if status == "extra_in_2b":
        return 0.0

    if status == "matched" and risk_level in ("High", "Medium"):
        return pr_itc

    return 0.0


def filter_exception_rows(df):
    """
    Return a copy of df containing only rows that require CA attention.

    Exceptions are rows where:
      - The invoice did not reconcile cleanly (status != 'matched'), OR
      - The supplier carries Medium or High risk (even if the invoice matched)

    Rows that are both 'matched' and 'Low' risk are considered clean and excluded.
    """
    mask = (df["status"] != "matched") | (df["supplier_risk_level"] != "Low")
    return df[mask].copy()


def get_suggested_ca_action(row):
    """
    Return a one-line action string for the CA based on the row's risk signals.

    Priority order matters — the first matching condition wins.
    """
    status = row.get("status", "")
    risk_level = row.get("supplier_risk_level", "Low")

    if status == "extra_in_2b":
        return "Check whether invoice is missing from books"

    if risk_level == "High":
        return "Review before filing"

    if risk_level == "Medium":
        return "Verify supporting documents"

    if status != "matched":
        return "Check reconciliation difference"

    return "No action required"


def prepare_exception_report(df):
    """
    Build a CA-readable exception report DataFrame from the enriched result.

    Steps:
      1. Filter to exception rows only
      2. Calculate ITC at risk per row
      3. Add suggested CA action per row
      4. Rename columns to human-readable headers
      5. Return the final DataFrame with 12 CA-readable columns

    Returns an empty DataFrame (with headers) if no exceptions are found.
    """
    filtered = filter_exception_rows(df)

    if filtered.empty:
        # Return a properly structured empty report — don't crash
        return pd.DataFrame(columns=list(COLUMN_RENAME.values()) + ["CA Review Status", "CA Remarks"])

    filtered = filtered.copy()
    filtered["itc_at_risk"] = filtered.apply(calculate_itc_at_risk, axis=1)
    filtered["suggested_ca_action"] = filtered.apply(get_suggested_ca_action, axis=1)

    # Select and rename to CA-readable columns in the defined order
    internal_cols = list(COLUMN_RENAME.keys())
    report = filtered[internal_cols].rename(columns=COLUMN_RENAME)

    # Add fillable CA review columns — the CA fills these in the Excel file
    report["CA Review Status"] = "Pending"
    report["CA Remarks"] = ""

    return report


def generate_report_filename():
    """
    Return a dated filename for the exception report.
    The caller is responsible for prepending the output directory.

    Example: "exception_report_20240516.xlsx"
    """
    today = datetime.date.today().strftime("%Y%m%d")
    return f"exception_report_{today}.xlsx"


def write_exception_report(df, output_path):
    """
    Write the exception report DataFrame to a formatted Excel file.

    Args:
        df: prepared exception report DataFrame (CA-readable column names)
        output_path: pathlib.Path — full path including filename

    Formatting applied:
      - Bold header row
      - Frozen top row
      - Autofilter on headers
      - Reasonable column widths
      - Row background colour by Supplier Risk Level (red=High, amber=Medium)
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Exception Report", index=False)
        ws = writer.sheets["Exception Report"]

        # Bold header row
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Freeze the header so it stays visible when scrolling
        ws.freeze_panes = "A2"

        # Autofilter on all columns
        ws.auto_filter.ref = ws.dimensions

        # Set column widths
        for col_idx, col_name in enumerate(df.columns, start=1):
            width = COLUMN_WIDTHS.get(col_name, 18)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

        # Find the "Supplier Risk Level" column index dynamically
        headers = [cell.value for cell in ws[1]]
        if "Supplier Risk Level" not in headers:
            return  # no risk column — skip row colouring

        risk_col_idx = headers.index("Supplier Risk Level") + 1  # 1-based

        # Colour data rows by risk level
        for row_idx in range(2, ws.max_row + 1):
            risk_value = ws.cell(row=row_idx, column=risk_col_idx).value
            if risk_value == "High":
                fill = HIGH_FILL
            elif risk_value == "Medium":
                fill = MEDIUM_FILL
            else:
                continue  # Low — no fill

            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill
